import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl import load_workbook
import shutil
import os
import zipfile
import re
import datetime
import json 

try:
    import win32com.client as win32
    PDF_SUPORTE = True
except ImportError:
    PDF_SUPORTE = False

# Para imagens na UI (capa do app)
try:
    from PIL import Image, ImageTk
    PIL_SUPORTE = True
except ImportError:
    PIL_SUPORTE = False
    print("Aviso: Pillow não instalado (pip install Pillow). Imagem de capa será texto fallback.")

class TextWithPlaceholder(tk.Text):
    def __init__(self, master=None, placeholder="PLACEHOLDER", color='#7f8c8d', *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self.cget('fg')
        self.bind("<FocusIn>", self._clear_placeholder)
        self.bind("<FocusOut>", self._add_placeholder)
        self._add_placeholder()

    def _clear_placeholder(self, event=None):
        content = self.get("1.0", tk.END).strip()
        if content == self.placeholder and self.cget('fg') == self.placeholder_color:
            self.delete("1.0", tk.END)
            self.config(fg=self.default_fg_color)

    def _add_placeholder(self, event=None):
        if not self.get("1.0", tk.END).strip():
            self.delete("1.0", tk.END)
            self.insert("1.0", self.placeholder)
            self.config(fg=self.placeholder_color)
            self.update_idletasks()

    def get_value(self):
        val = self.get("1.0", tk.END).strip()
        if val == self.placeholder and self.cget('fg') == self.placeholder_color:
            return ""
        return val

    def set_placeholder(self, text):
        self.placeholder = text if text else "PLACEHOLDER"
        self.delete("1.0", tk.END)
        self._add_placeholder()
        self.event_generate("<FocusOut>")
        self.update()

class EntryWithPlaceholder(tk.Entry):
    def __init__(self, master=None, placeholder="PLACEHOLDER", color='#7f8c8d', *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self['fg']
        self.bind("<FocusIn>", self._clear_placeholder)
        self.bind("<FocusOut>", self._add_placeholder)
        self.bind("<KeyRelease>", self._auto_scroll)
        self._add_placeholder()

    def _clear_placeholder(self, event=None):
        if self['fg'] == self.placeholder_color and self.get() == self.placeholder:
            self.delete(0, tk.END)
            self['fg'] = self.default_fg_color

    def _add_placeholder(self, event=None):
        if not self.get():
            self.delete(0, tk.END)
            self.insert(0, self.placeholder)
            self['fg'] = self.placeholder_color
            self.update_idletasks()

    def _auto_scroll(self, event=None):
        if len(self.get()) > self.cget('width'):
            self.xview_moveto(1.0)

    def get_value(self):
        val = self.get()
        if val == self.placeholder and self['fg'] == self.placeholder_color:
            return ""
        return val

    def set_placeholder(self, text):
        self.placeholder = text if text else "PLACEHOLDER"
        self.delete(0, tk.END)
        self._add_placeholder()
        self.event_generate("<FocusOut>")
        self.update()

    def set_real_value(self, text):
        """Define valor real (preto, não placeholder)."""
        self.delete(0, tk.END)
        self.insert(0, text)
        self['fg'] = self.default_fg_color
        self.update_idletasks()

class Automatizador:
    def __init__(self, root):
        self.root = root
        self.root.title("Automatizador de Orçamentos Excel")
        self.root.geometry("1000x750")
        self.root.minsize(850, 650)
        self.root.configure(bg='#f0f0f0')  # Fundo claro suave
        self.root.resizable(True, True)

        # Estilos ttk refinados para UI mais bonita
        self.style = ttk.Style()
        self.style.theme_use('clam')  # Tema moderno
        self.style.configure('Title.TLabel', font=('Arial', 12, 'bold'), foreground='#2c3e50')
        self.style.configure('Header.TLabel', font=('Arial', 14, 'bold'), foreground='#2c3e50')
        self.style.configure('TButton', font=('Arial', 10, 'bold'), padding=(15, 8))
        self.style.configure('Select.TButton', background='#3498db', foreground='white')
        self.style.configure('Create.TButton', background='#27ae60', foreground='white')
        self.style.map('Select.TButton', background=[('active', '#2980b9')], foreground=[('active', 'white')])
        self.style.map('Create.TButton', background=[('active', '#229954')], foreground=[('active', 'white')])
        self.style.configure('TScrollbar', gripcount=0, borderwidth=1)  # Scrollbar mais fina

        # Tooltip simples (popup label)
        self.tooltip = None

        # Configuração da imagem de capa (logo no topo, com sombra)
        self.logo_path = "logo.png"
        self.logo_image = None
        self.logo_label = None
        if PIL_SUPORTE:
            if os.path.exists(self.logo_path):
                try:
                    pil_image = Image.open(self.logo_path)
                    pil_image = pil_image.resize((280, 120), Image.Resampling.LANCZOS)
                    self.logo_image = ImageTk.PhotoImage(pil_image)
                    self.log("Imagem de capa carregada: logo.png (redimensionada para 280x120).")
                    # Ícone da janela
                    root_icon = Image.open(self.logo_path)
                    root_icon = root_icon.resize((32, 32), Image.Resampling.LANCZOS)
                    root_icon_tk = ImageTk.PhotoImage(root_icon)
                    self.root.iconphoto(False, root_icon_tk)
                except Exception as img_err:
                    self.log(f"Erro ao carregar imagem de capa: {img_err}. Usando texto fallback.")
                    self.logo_image = None
            else:
                self.log(f"Imagem de capa não encontrada: {self.logo_path}. Usando texto fallback.")
                self.logo_image = None
        else:
            self.log("Pillow não disponível. Usando texto fallback para capa.")

        # Persistência do número (config.json)
        self.config_file = "config.json"
        self.numero_orcamento = tk.StringVar(value="1")
        self.carregar_numero_config()

        self.pasta_selecionada = tk.StringVar()
        self.arquivo_selecionado = tk.StringVar()

        self.campos_fixos = [
            {'titulo': 'Cliente', 'valor': tk.StringVar(value='Nome do Cliente'), 'celula_fixa': 'B6', 'entry': None, 'tooltip': 'Digite o nome do cliente aqui'},
            {'titulo': 'Endereço', 'valor': tk.StringVar(value='Endereço do Cliente'), 'celula_fixa': 'B7', 'entry': None, 'tooltip': 'Digite o endereço completo'},
            {'titulo': 'CNPJ', 'valor': tk.StringVar(value='00.000.000/0000-00'), 'celula_fixa': 'B8', 'entry': None, 'tooltip': 'Digite o CNPJ no formato XX.XXX.XXX/XXXX-XX'},
            {'titulo': 'Telefone', 'valor': tk.StringVar(value='(00) 0000-0000'), 'celula_fixa': 'B9', 'entry': None, 'tooltip': 'Digite o telefone no formato (XX) XXXX-XXXX'}
        ]

        self.campos_adicionais = [
            {'titulo': 'Prazo de Entrega (B36)', 'celula': 'B36', 'placeholder_default': 'Prazo de entrega: ', 'entry': None, 'tooltip': 'Edite o prazo carregado do modelo (preto se presente)'},
            {'titulo': 'Forma de Pagamento (B37)', 'celula': 'B37', 'placeholder_default': 'Forma de pagamento: ', 'entry': None, 'tooltip': 'Edite a forma de pagamento (preto se presente no modelo)'},
            {'titulo': 'Condições (B39)', 'celula': 'B39', 'placeholder_default': 'Na entrega: ', 'entry': None, 'tooltip': 'Edite as condições (preto se presente no modelo)'}
        ]

        self.itens_widgets = []
        self.label_total_preview = None  # Para preview F35

        self.main_container = tk.Frame(root, bg='#f0f0f0')
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        self.canvas_principal = tk.Canvas(self.main_container, bg='#f0f0f0', highlightthickness=0)
        self.scrollbar_principal = ttk.Scrollbar(self.main_container, orient="vertical", command=self.canvas_principal.yview, style='TScrollbar')
        self.scrollable_frame_principal = tk.Frame(self.canvas_principal, bg='#f0f0f0')

        self.scrollable_frame_principal.bind(
            "<Configure>",
            lambda e: self.canvas_principal.configure(scrollregion=self.canvas_principal.bbox("all"))
        )

        self.canvas_principal.create_window((0, 0), window=self.scrollable_frame_principal, anchor="nw")
        self.canvas_principal.configure(yscrollcommand=self.scrollbar_principal.set)

        self.canvas_principal.pack(side="left", fill="both", expand=True)
        self.scrollbar_principal.pack(side="right", fill="y")

        self.main_container.rowconfigure(0, weight=1)
        self.main_container.columnconfigure(0, weight=1)

        self.configurar_interface()

        # Bind para preview total (atualiza ao digitar em totais de itens)
        self.bind_preview_total()

    def carregar_numero_config(self):
        """Carrega próximo número de config.json (fallback se sem modelo)."""
        try:
            self.log("Iniciando carregamento de config.json (fallback)...")
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    proximo = int(config.get('proximo_numero', 1))
                    self.numero_orcamento.set(str(proximo))
                    self.log(f"Número carregado de config.json (fallback): {proximo}")
                    print(f"DEBUG FALLBACK: Carregado {proximo} de JSON (será sobrescrito por A5 se modelo carregado)")
            else:
                self.log("config.json não encontrado. Fallback para 1.")
                self.numero_orcamento.set("1")
                print("DEBUG FALLBACK: Sem JSON, usando 1")
        except Exception as e:
            self.log(f"Erro config.json (fallback): {e}. Usando 1.")
            self.numero_orcamento.set("1")
        
        # Force update
        self.root.after(0, lambda: self.entry_numero.update_idletasks())

    def salvar_numero_config(self, novo_numero):
        """Salva próximo número em config.json."""
        try:
            self.log(f"Tentando salvar {novo_numero} em config.json...")
            config = {'proximo_numero': novo_numero}
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            self.log(f"Número salvo em config.json: {novo_numero}")
            print(f"DEBUG JSON: Salvo {novo_numero} em {self.config_file} (conteúdo: {config})")  # Debug extra
            
            # Verifica se salvou
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    verificado = json.load(f)
                    print(f"DEBUG JSON: Verificação pós-save: {verificado}")
            else:
                print("DEBUG JSON: ERRO - Arquivo não criado após save!")
        except Exception as e:
            self.log(f"Erro ao salvar config.json: {e}. Número não persistido (verifique permissões/pasta).")
            print(f"DEBUG JSON: Exceção save: {e} (pasta atual: {os.getcwd()})")

    def log(self, mensagem):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {mensagem}")

    def mostrar_tooltip(self, widget, texto, event):
        """Mostra tooltip simples (popup label)."""
        if self.tooltip:
            self.tooltip.destroy()
        x = event.x_root + 10
        y = event.y_root + 10
        self.tooltip = tk.Toplevel(self.root)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip, text=texto, background='#ffffe0', relief='solid', borderwidth=1, padx=5, pady=2, font=('Arial', 9))
        label.pack()
        self.root.after(2000, self.tooltip.destroy)  # Desaparece em 2s

    def configurar_interface(self):
        row = 0

        # Header aprimorado com gradiente simulado e sombra
        header_frame = tk.Frame(self.scrollable_frame_principal, height=160, bg='white', relief=tk.SUNKEN, bd=3)
        header_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(0, 25))
        header_frame.rowconfigure(0, weight=1)
        header_frame.columnconfigure(0, weight=1)

        # Canvas para fundo gradiente simples (branco para cinza claro)
        canvas_header = tk.Canvas(header_frame, height=20, bg='#e8f4fd', highlightthickness=0)
        canvas_header.pack(fill=tk.X)
        canvas_header.create_rectangle(0, 0, 1000, 20, fill='#d5f4e6', outline='')  # Gradiente simulado

        # Sub-frame para logo e título
        logo_subframe = tk.Frame(header_frame, bg='white')
        logo_subframe.pack(pady=10)

        if self.logo_image:
            self.logo_label = tk.Label(logo_subframe, image=self.logo_image, bg='white', relief=tk.SUNKEN, bd=2)
            self.logo_label.pack(padx=20)
            # Tooltip no logo (se presente)
            self.logo_label.bind("<Enter>", lambda e: self.mostrar_tooltip(self.logo_label, "Logo da Empresa", e))
            self.logo_label.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)
        else:
            # Fallback: Texto centralizado com estilo bold/italic
            fallback_frame = tk.Frame(logo_subframe, bg='white')
            fallback_frame.pack()
            fallback_label = tk.Label(fallback_frame, text="Automatizador de Orçamentos Excel", font=('Arial', 20, 'bold'), bg='white', fg='#2c3e50')
            fallback_label.pack(pady=(0, 5))
            subtitle_label = tk.Label(fallback_frame, text="Gere cópias .xlsx e PDF", font=('Arial', 12, 'italic'), bg='white', fg='#7f8c8d')
            subtitle_label.pack()
            # Tooltip no fallback
            fallback_label.bind("<Enter>", lambda e: self.mostrar_tooltip(fallback_label, "Sistema para orçamentos automatizados", e))
            fallback_label.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        # Título principal
        titulo_label = tk.Label(header_frame, text="🚀 Automatizador de Orçamentos Excel", font=('Arial', 16, 'bold'), bg='white', fg='#27ae60')
        titulo_label.pack(pady=5)
        # Tooltip no título
        titulo_label.bind("<Enter>", lambda e: self.mostrar_tooltip(titulo_label, "Carregue modelo, edite e gere cópias com A5 +1 automático", e))
        titulo_label.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Seção Seleção de Pasta e Arquivo (LabelFrame com tooltips)
        pasta_frame = ttk.LabelFrame(self.scrollable_frame_principal, text="📁 Seleção de Pasta e Arquivo Modelo", padding=15)
        pasta_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=10)
        pasta_frame.columnconfigure(1, weight=1)

        tk.Label(pasta_frame, text="Pasta dos Arquivos:", font=('Arial', 10), bg='#f0f0f0').grid(row=0, column=0, sticky="e", padx=5, pady=8)
        pasta_entry = tk.Entry(pasta_frame, textvariable=self.pasta_selecionada, width=50, relief='solid', bd=1, font=('Arial', 9))
        pasta_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=8)
        select_pasta_btn = ttk.Button(pasta_frame, text="📁 Selecionar Pasta", style='Select.TButton', command=lambda: self.selecionar_pasta(pasta_entry))
        select_pasta_btn.grid(row=0, column=2, padx=5, pady=8)
        # Hover e tooltip para botão pasta
        select_pasta_btn.bind("<Enter>", lambda e: self.mostrar_tooltip(select_pasta_btn, "Selecione a pasta com o arquivo modelo", e))
        select_pasta_btn.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        tk.Label(pasta_frame, text="Arquivo Modelo (.xlsx):", font=('Arial', 10), bg='#f0f0f0').grid(row=1, column=0, sticky="e", padx=5, pady=8)
        arquivo_entry = tk.Entry(pasta_frame, textvariable=self.arquivo_selecionado, width=50, relief='solid', bd=1, font=('Arial', 9))
        arquivo_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=8)
        select_arquivo_btn = ttk.Button(pasta_frame, text="📄 Selecionar Arquivo", style='Select.TButton', command=lambda: self.selecionar_arquivo(arquivo_entry))
        select_arquivo_btn.grid(row=1, column=2, padx=5, pady=8)
        # Hover e tooltip para botão arquivo
        select_arquivo_btn.bind("<Enter>", lambda e: self.mostrar_tooltip(select_arquivo_btn, "Carregue valores do modelo (A5 para auto-incremento, B36 preto se presente)", e))
        select_arquivo_btn.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Separador mais grosso (corrigido: Separator)
        ttk.Separator(self.scrollable_frame_principal, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky="ew", pady=20)
        row += 1

        # Número do Orçamento (com persistência e base em A5)
        numero_frame = tk.Frame(self.scrollable_frame_principal, bg='#f0f0f0')
        numero_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        tk.Label(numero_frame, text="Número do Orçamento (auto-incrementa baseado em A5 do modelo):", font=('Arial', 10), bg='#f0f0f0').grid(row=0, column=0, sticky="w", padx=5, pady=8)
        self.entry_numero = tk.Entry(numero_frame, textvariable=self.numero_orcamento, width=10, relief='solid', bd=1, font=('Arial', 9))
        self.entry_numero.grid(row=0, column=1, sticky="w", padx=5, pady=8)
        # Tooltip para número
        self.entry_numero.bind("<Enter>", lambda e: self.mostrar_tooltip(self.entry_numero, "Carregado de A5 do modelo +1 (persistido em config.json)", e))
        self.entry_numero.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)
        row += 1

        # Seção Dados do Cliente (LabelFrame com tooltips)
        cliente_frame = ttk.LabelFrame(self.scrollable_frame_principal, text="👤 Dados do Cliente", padding=15)
        cliente_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        cliente_frame.columnconfigure(1, weight=1)

        tk.Label(cliente_frame, text="Título", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').grid(row=0, column=0, sticky="w", padx=5, pady=8)
        tk.Label(cliente_frame, text="Valor", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').grid(row=0, column=1, sticky="w", padx=5, pady=8)

        for idx, campo in enumerate(self.campos_fixos):
            row_campo = idx + 1
            lbl = tk.Label(cliente_frame, text=campo['titulo'], font=('Arial', 9), bg='#f0f0f0')
            lbl.grid(row=row_campo, column=0, sticky="w", padx=5, pady=6)
            entry = EntryWithPlaceholder(cliente_frame, placeholder=campo['valor'].get(), width=50, relief='solid', bd=1, font=('Arial', 9))
            entry.grid(row=row_campo, column=1, sticky="ew", padx=5, pady=6)
            campo['entry'] = entry
            # Tooltip para cada campo
            entry.bind("<Enter>", lambda e, tt=campo['tooltip']: self.mostrar_tooltip(entry, tt, e))
            entry.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Separador (corrigido: Separator)
        ttk.Separator(self.scrollable_frame_principal, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky="ew", pady=20)
        row += 1

        # Seção Informações Adicionais (LabelFrame com tooltips)
        adicionais_frame = ttk.LabelFrame(self.scrollable_frame_principal, text="ℹ️ Informações Adicionais (B36, B37, B39 – preto se presentes no modelo)", padding=15)
        adicionais_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        adicionais_frame.columnconfigure(1, weight=1)

        tk.Label(adicionais_frame, text="Título", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').grid(row=0, column=0, sticky="w", padx=5, pady=8)
        tk.Label(adicionais_frame, text="Valor", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').grid(row=0, column=1, sticky="w", padx=5, pady=8)

        for idx, campo in enumerate(self.campos_adicionais):
            row_campo = idx + 1
            lbl = tk.Label(adicionais_frame, text=campo['titulo'], font=('Arial', 9), bg='#f0f0f0')
            lbl.grid(row=row_campo, column=0, sticky="w", padx=5, pady=6)
            placeholder_init = campo['placeholder_default']
            entry = EntryWithPlaceholder(adicionais_frame, placeholder=placeholder_init, width=50, relief='solid', bd=1, font=('Arial', 9))
            entry.grid(row=row_campo, column=1, sticky="ew", padx=5, pady=6)
            campo['entry'] = entry
            # Tooltip para cada campo adicional
            entry.bind("<Enter>", lambda e, tt=campo['tooltip']: self.mostrar_tooltip(entry, tt, e))
            entry.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Separador (corrigido: Separator)
        ttk.Separator(self.scrollable_frame_principal, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky="ew", pady=20)
        row += 1

        # Seção Itens do Orçamento (LabelFrame com tabela e preview total)
        itens_frame = ttk.LabelFrame(self.scrollable_frame_principal, text="📋 Itens do Orçamento (A13-F34 – cinza do modelo; preview total abaixo)", padding=15)
        itens_frame.grid(row=row, column=0, columnspan=6, sticky="nsew", pady=5)
        itens_frame.columnconfigure(0, weight=1)
        itens_frame.rowconfigure(1, weight=1)

        # Títulos das colunas (estilizados com fundo azul escuro)
        colunas_frame = tk.Frame(itens_frame, bg='#f0f0f0', relief=tk.RIDGE, bd=2)
        colunas_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        colunas_config = [
            ("Item (Col A)", 10),
            ("Descrição (Col B)", 35),
            ("Quantidade (Col C)", 12),
            ("UND (Col D)", 8),
            ("Valor Uni (Col E)", 15),
            ("Total (Col F)", 15)
        ]
        for col_idx, (texto, width) in enumerate(colunas_config):
            lbl = tk.Label(colunas_frame, text=texto, font=('Arial', 9, 'bold'), width=width, anchor='w', bg='#34495e', fg='white', relief=tk.SOLID, bd=1)
            lbl.grid(row=0, column=col_idx, sticky="ew", padx=1)

        # Container para itens com canvas e scrollbar
        itens_container = tk.Frame(itens_frame, bg='#f0f0f0')
        itens_container.grid(row=1, column=0, sticky="nsew")
        itens_container.rowconfigure(0, weight=1)
        itens_container.columnconfigure(0, weight=1)

        self.canvas_itens = tk.Canvas(itens_container, bg='#f0f0f0', highlightthickness=0, height=300)
        self.scrollbar_itens = ttk.Scrollbar(itens_container, orient="vertical", command=self.canvas_itens.yview, style='TScrollbar')
        self.scrollable_frame_itens = tk.Frame(self.canvas_itens, bg='#f0f0f0')

        for col_idx, (_, width) in enumerate(colunas_config):
            self.scrollable_frame_itens.columnconfigure(col_idx, minsize=width * 7)

        self.scrollable_frame_itens.bind(
            "<Configure>",
            lambda e: self.canvas_itens.configure(scrollregion=self.canvas_itens.bbox("all"))
        )

        self.canvas_itens.create_window((0, 0), window=self.scrollable_frame_itens, anchor="nw")
        self.canvas_itens.configure(yscrollcommand=self.scrollbar_itens.set)

        self.canvas_itens.pack(side="left", fill="both", expand=True)
        self.scrollbar_itens.pack(side="right", fill="y")

        self.pre_carregar_itens()  # Carrega itens com estilos e binds

        # Label Preview Total (só-leitura, atualiza em tempo real)
        preview_frame = tk.Frame(itens_frame, bg='#f0f0f0')
        preview_frame.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        self.label_total_preview = tk.Label(preview_frame, text="Total Estimado (F35): R$ 0,00", font=('Arial', 11, 'bold'), bg='#f0f0f0', fg='#27ae60', anchor='e')
        self.label_total_preview.pack(anchor='e')
        # Tooltip para preview
        self.label_total_preview.bind("<Enter>", lambda e: self.mostrar_tooltip(self.label_total_preview, "Soma automática dos totais editados (F13-F34, formato BR)", e))
        self.label_total_preview.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Separador final (corrigido: Separator)
        ttk.Separator(self.scrollable_frame_principal, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky="ew", pady=20)
        row += 1

        # Botão Principal (estilizado com hover e loading)
        botao_frame = tk.Frame(self.scrollable_frame_principal, bg='#f0f0f0')
        botao_frame.grid(row=row, column=0, columnspan=3, pady=20)
        self.botao_criar = ttk.Button(botao_frame, text="🚀 Criar Cópia do Orçamento e PDF" if PDF_SUPORTE else "💾 Criar Cópia do Orçamento (.xlsx apenas)", style='Create.TButton', command=self.executar)
        self.botao_criar.pack(pady=10)
        # Hover e tooltip para botão criar
        self.botao_criar.bind("<Enter>", lambda e: self.mostrar_tooltip(self.botao_criar, "Gera cópia com valores editados, A5 incrementado, imagens preservadas e PDF (se Windows)", e))
        self.botao_criar.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        row += 1

        # Footer com versão e crédito (italic 💙)
        footer_frame = tk.Frame(self.scrollable_frame_principal, bg='#f0f0f0')
        footer_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=(20, 0))
        footer_label = tk.Label(footer_frame, text="Versão 2.0 - Automatizador com A5 Auto-Incremento Baseado no Modelo 💙", font=('Arial', 9, 'italic'), bg='#f0f0f0', fg='#3498db', anchor='center')
        footer_label.pack()
        # Tooltip no footer
        footer_label.bind("<Enter>", lambda e: self.mostrar_tooltip(footer_label, "App desenvolvido para orçamentos Excel - Testado com Python 3.8+", e))
        footer_label.bind("<Leave>", lambda e: self.tooltip.destroy() if self.tooltip else None)

        # Configurações de grid para responsividade
        self.scrollable_frame_principal.columnconfigure(0, weight=1)
        for i in range(row + 1):
            self.scrollable_frame_principal.rowconfigure(i, weight=0)
        self.scrollable_frame_principal.rowconfigure(row + 1, weight=1)  # Espaço extra no final

        self.log("Interface carregada (UI mais bonita: gradiente header, tooltips, preview total live, hover botões, LabelFrames com emojis).")

    def bind_preview_total(self):
        """Bind KeyRelease nos totais de itens para preview live."""
        for item in self.itens_widgets:
            total_entry = item['total']
            total_entry.bind("<KeyRelease>", self.atualizar_preview_total)

    def atualizar_preview_total(self, event=None):
        """Atualiza label preview com soma dos totais editados (F13-F34, formato BR)."""
        total_soma = 0.0
        try:
            for i, item in enumerate(self.itens_widgets):
                valor_total_str = item['total'].get_value()
                if valor_total_str and valor_total_str != "0,00":
                    valor_limpo = valor_total_str.replace('.', '').replace(',', '.')
                    try:
                        total_num = float(valor_limpo)
                        total_soma += total_num
                    except ValueError:
                        pass  # Ignora inválidos
            # Formato BR (pontos milhar, vírgula decimal)
            total_formatado = f"{total_soma:,.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
            self.label_total_preview.config(text=f"Total Estimado (F35): R$ {total_formatado}")
            self.log(f"Preview total atualizado: R$ {total_formatado}")
        except Exception as e:
            self.log(f"Erro no preview total: {e}")
            self.label_total_preview.config(text="Total Estimado (F35): R$ 0,00")

    def pre_carregar_itens(self):
        """Pré-carrega 22 itens com estilos bonitos (cores alternadas, bordas) e binds para preview."""
        colunas_placeholders_iniciais = [
            "Item 1", "Digite a descrição do item", "0", "UND", "0,00", "0,00"
        ]
        widths = [10, 35, 12, 8, 15, 15]

        for i in range(22):  # Linhas 13-34
            row_item = i
            placeholder_item = f"Item {i+1}"

            # Cores alternadas suaves
            bg_color = '#f8f9fa' if i % 2 == 0 else '#ffffff'

            # Coluna A: Item
            item_entry = EntryWithPlaceholder(self.scrollable_frame_itens, placeholder=placeholder_item, width=widths[0], relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            item_entry.grid(row=row_item, column=0, padx=1, pady=1, sticky="ew")

            # Coluna B: Descrição (Text)
            desc_text = TextWithPlaceholder(self.scrollable_frame_itens, placeholder=colunas_placeholders_iniciais[1], width=widths[1], height=1, wrap=tk.NONE, relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            desc_text.grid(row=row_item, column=1, padx=1, pady=1, sticky="ew")

            # Colunas C-F: Entries
            quant_entry = EntryWithPlaceholder(self.scrollable_frame_itens, placeholder=colunas_placeholders_iniciais[2], width=widths[2], relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            quant_entry.grid(row=row_item, column=2, padx=1, pady=1, sticky="ew")

            und_entry = EntryWithPlaceholder(self.scrollable_frame_itens, placeholder=colunas_placeholders_iniciais[3], width=widths[3], relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            und_entry.grid(row=row_item, column=3, padx=1, pady=1, sticky="ew")

            vlr_uni_entry = EntryWithPlaceholder(self.scrollable_frame_itens, placeholder=colunas_placeholders_iniciais[4], width=widths[4], relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            vlr_uni_entry.grid(row=row_item, column=4, padx=1, pady=1, sticky="ew")

            total_entry = EntryWithPlaceholder(self.scrollable_frame_itens, placeholder=colunas_placeholders_iniciais[5], width=widths[5], relief='solid', bd=1, bg=bg_color, font=('Arial', 9))
            total_entry.grid(row=row_item, column=5, padx=1, pady=1, sticky="ew")

            # Armazena widgets
            item_widgets = {
                'item': item_entry, 'desc': desc_text, 'quant': quant_entry,
                'und': und_entry, 'vlr_uni': vlr_uni_entry, 'total': total_entry
            }
            self.itens_widgets.append(item_widgets)

        self.log("Itens pré-carregados (22 linhas com estilos: alternadas '#f8f9fa'/'#ffffff', bordas, binds para preview total).")

    def selecionar_pasta(self, entry):
        pasta = filedialog.askdirectory()
        if pasta:
            entry.delete(0, tk.END)
            entry.insert(0, pasta)
            self.log(f"Pasta selecionada: {pasta}")

    def selecionar_arquivo(self, entry):
        pasta = self.pasta_selecionada.get()
        if not pasta:
            messagebox.showwarning("Aviso", "Selecione a pasta primeiro.")
            return
        arquivo = filedialog.askopenfilename(initialdir=pasta, title="Selecione o arquivo modelo .xlsx", filetypes=[("Excel files", "*.xlsx")])
        if arquivo:
            nome_arquivo = os.path.basename(arquivo)
            entry.delete(0, tk.END)
            entry.insert(0, nome_arquivo)
            self.arquivo_selecionado.set(nome_arquivo)
            self.log(f"Arquivo selecionado: {nome_arquivo}. Carregando valores do modelo...")
            self.carregar_valores_modelo(arquivo)

    def obter_valor_real_celula(self, ws, celula):
        """Obtém valor real da célula (ignora fórmulas vazias)."""
        try:
            cel = ws[celula]
            if cel is None:
                return ""
            valor = cel.value
            if valor is None:
                return ""
            valor_str = str(valor).strip()
            if not valor_str:
                return ""
            return valor_str
        except Exception as e:
            self.log(f"Erro ao obter valor de {celula}: {e}")
            return ""

    def extrair_numero_a5(self, valor_a5):
        """Extrai número de A5 (ex: 'N° do Orçamento 5' → 5). Fallback 1 se inválido."""
        if not valor_a5:
            self.log("A5 vazio. Fallback para 1.")
            return 1
        try:
            # Regex para extrair dígitos no final (flexível: 'Orçamento 5', 'N° 10', etc.)
            match = re.search(r'\d+', valor_a5)
            if match:
                numero_extraido = int(match.group())
                self.log(f"Número extraído de A5 '{valor_a5}': {numero_extraido}")
                return numero_extraido
            else:
                self.log(f"A5 sem número válido '{valor_a5}'. Fallback para 1.")
                return 1
        except ValueError:
            self.log(f"Erro extrair número de A5 '{valor_a5}'. Fallback para 1.")
            return 1

    def carregar_valores_modelo(self, caminho_arquivo):
        """Carrega valores: preto para B36/B37/B39; cinza para fixos/itens. Extrai A5 para auto-incremento."""
        try:
            wb = load_workbook(caminho_arquivo)
            ws = wb.active

            # NOVA PARTE: Extrai número de A5 para auto-incremento
            valor_a5 = self.obter_valor_real_celula(ws, 'A5')
            numero_a5 = self.extrair_numero_a5(valor_a5)
            
            # Carrega de JSON (se existir) e usa o MAIOR (persistência + modelo)
            try:
                if os.path.exists(self.config_file):
                    with open(self.config_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                        numero_json = int(config.get('proximo_numero', 1))
                        proximo_numero = max(numero_a5 + 1, numero_json)  # Usa o maior
                        self.log(f"A5: {numero_a5} (JSON: {numero_json}) → Próximo: {proximo_numero}")
                else:
                    proximo_numero = numero_a5 + 1
                    self.log(f"A5: {numero_a5} (sem JSON) → Próximo: {proximo_numero}")
            except Exception as json_err:
                proximo_numero = numero_a5 + 1
                self.log(f"Erro JSON em carregamento: {json_err}. Usando A5 +1: {proximo_numero}")
            
            self.numero_orcamento.set(str(proximo_numero))
            self.salvar_numero_config(proximo_numero)  # Salva/atualiza JSON
            print(f"DEBUG A5: Extraído {numero_a5} de '{valor_a5}' → UI/JSON definido para {proximo_numero}")

            # Force update visual
            self.entry_numero.delete(0, tk.END)
            self.entry_numero.insert(0, str(proximo_numero))
            self.entry_numero.update_idletasks()

            # Campos fixos B6-B9: placeholders cinza
            for campo in self.campos_fixos:
                valor_modelo = self.obter_valor_real_celula(ws, campo['celula_fixa'])
                if valor_modelo:
                    entry = campo['entry']
                    entry.set_placeholder(valor_modelo)
                    self.log(f"Campo fixo '{campo['titulo']}' carregado como placeholder cinza: {valor_modelo}")
                else:
                    self.log(f"Campo fixo '{campo['titulo']}' vazio, mantendo placeholder cinza.")

            # Campos adicionais B36/B37/B39: preto se presente, cinza se vazio
            for campo in self.campos_adicionais:
                valor_modelo = self.obter_valor_real_celula(ws, campo['celula'])
                entry = campo['entry']
                if valor_modelo:
                    entry.set_real_value(valor_modelo)  # Preto
                    self.log(f"Campo adicional '{campo['titulo']}' carregado como real preto: {valor_modelo}")
                else:
                    entry.set_placeholder(campo['placeholder_default'])  # Cinza
                    self.log(f"Campo adicional '{campo['titulo']}' vazio, placeholder cinza.")

            # Itens A13-F34: placeholders cinza
            for i, item in enumerate(self.itens_widgets):
                linha = 13 + i
                if linha > 34:
                    break
                colunas = [('item', 'A'), ('desc', 'B'), ('quant', 'C'), ('und', 'D'), ('vlr_uni', 'E'), ('total', 'F')]
                for key, col in colunas:
                    celula = f"{col}{linha}"
                    valor_modelo = self.obter_valor_real_celula(ws, celula)
                    widget = item[key]
                    if valor_modelo:
                        if isinstance(widget, TextWithPlaceholder):
                            widget.set_placeholder(valor_modelo)
                        else:
                            widget.set_placeholder(valor_modelo)
                        self.log(f"Item {i+1} {key.upper()} carregado como placeholder cinza em {celula}: {valor_modelo}")
                    else:
                        self.log(f"Item {i+1} {key.upper()} vazio, mantendo placeholder cinza.")

            wb.close()
            self.log("Valores carregados (preto B36/B37/B39 se presentes; cinza fixos/itens). A5 processado para auto-incremento.")

            # Atualiza UI e preview após carregamento
            self.root.after(100, self.atualizar_campos_automaticos)
            self.root.after(300, self.atualizar_preview_total)  # Preview após load

        except Exception as e:
            self.log(f"Erro ao carregar modelo: {e}")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}")
            messagebox.showerror("Erro", f"Falha ao carregar .xlsx: {e}")

    def atualizar_campos_automaticos(self):
        """Atualiza placeholders e UI (com delay para suavidade)."""
        try:
            for campo in self.campos_fixos + self.campos_adicionais:
                entry = campo['entry']
                entry.event_generate("<FocusOut>")
            for item in self.itens_widgets:
                for key in item:
                    widget = item[key]
                    widget.event_generate("<FocusOut>")
            self.canvas_itens.update_idletasks()
            self.canvas_itens.configure(scrollregion=self.canvas_itens.bbox("all"))
            self.log("UI atualizada após carregamento do modelo (placeholders aplicados, scroll ajustado).")
        except Exception as e:
            self.log(f"Erro atualizar campos automáticos: {e}")

    def forcar_atualizacao_ui(self):
        """Força update visual (mantém estilos)."""
        try:
            for campo in self.campos_fixos + self.campos_adicionais:
                entry = campo['entry']
                entry.update()
                entry.event_generate("<FocusOut>")
            
            for item in self.itens_widgets:
                for key in item:
                    widget = item[key]
                    widget.update()
                    widget.event_generate("<FocusOut>")
                    if hasattr(widget, 'see'):
                        widget.see(tk.END)
            
            self.canvas_principal.update_idletasks()
            self.canvas_principal.configure(scrollregion=self.canvas_principal.bbox("all"))
            self.canvas_itens.update_idletasks()
            self.canvas_itens.configure(scrollregion=self.canvas_itens.bbox("all"))
            
            self.atualizar_preview_total()  # Atualiza preview após UI
            
            self.log("UI forçada (estilos preservados, preview total atualizado).")
        except Exception as e:
            self.log(f"Erro forçar UI: {e}")

    def copiar_imagens_completo(self, ws_original, ws_copia):
        """Copia imagens usando openpyxl (preserva no Excel)."""
        imagens_detectadas = 0
        imagens_copiadas = 0
        try:
            if hasattr(ws_original, '_images') and ws_original._images:
                imagens_detectadas += len(ws_original._images)
            if hasattr(ws_original, '_drawings') and ws_original._drawings:
                imagens_detectadas += len(ws_original._drawings)
            self.log(f"{imagens_detectadas} imagem(ns)/drawing(s) detectada(s).")

            if imagens_detectadas == 0:
                self.log("Nenhuma imagem. Fallback ZIP se necessário.")
                return

            # Copia _images
            if hasattr(ws_original, '_images') and ws_original._images:
                for img in ws_original._images:
                    try:
                        nova_img = Image()
                        if hasattr(img, 'anchor'):
                            nova_img.anchor = img.anchor
                        if hasattr(img, 'width') and hasattr(img, 'height'):
                            nova_img.width = img.width
                            nova_img.height = img.height
                        if hasattr(img, '_data'):
                            nova_img._data = img._data
                        ws_copia.add_image(nova_img, str(img.anchor))
                        imagens_copiadas += 1
                    except Exception as img_err:
                        self.log(f"Aviso copiar imagem: {img_err}. Fallback ZIP.")
                self.log(f"{imagens_copiadas} imagem(ns) copiadas via _images.")

            # Copia _drawings
            drawings_copiados = 0
            if hasattr(ws_original, '_drawings') and ws_original._drawings:
                drawing_copia = SpreadsheetDrawing()
                for dr in ws_original._drawings:
                    try:
                        drawing_copia.add_drawing(dr)
                        drawings_copiados += 1
                    except Exception as dr_err:
                        self.log(f"Aviso copiar drawing: {dr_err}.")
                if drawings_copiados > 0:
                    ws_copia._drawings = [drawing_copia]
                    imagens_copiadas += drawings_copiados
                    self.log(f"{drawings_copiados} drawing(s) copiados.")

            self.log(f"Total {imagens_copiadas} imagens/desenhos copiadas (fallback ZIP se detectadas).")

        except Exception as e:
            self.log(f"Erro openpyxl imagens: {e}. Fallback ZIP.")

    def fallback_zip_imagens(self, arquivo_original, arquivo_copia):
        """Fallback: Re-insere mídia via ZIP."""
        midia_reinserida = 0
        try:
            if not os.path.exists(arquivo_original):
                self.log("Fallback ZIP: Original não encontrado.")
                return 0

            with zipfile.ZipFile(arquivo_original, 'r') as zip_orig:
                arquivos_media = [f for f in zip_orig.namelist() if f.startswith('xl/media/') and f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.emf'))]
                self.log(f"Fallback ZIP: {len(arquivos_media)} mídia(s) detectada(s).")

                if not arquivos_media:
                    return 0

                temp_dir = 'temp_media'
                os.makedirs(temp_dir, exist_ok=True)

                for media_file in arquivos_media:
                    zip_orig.extract(media_file, temp_dir)
                    self.log(f"Fallback ZIP: Extraído {os.path.basename(media_file)}.")

                with zipfile.ZipFile(arquivo_copia, 'a') as zip_copia:
                    for media_file in arquivos_media:
                        temp_path = os.path.join(temp_dir, media_file.replace('xl/media/', ''))
                        if os.path.exists(temp_path):
                            zip_copia.write(temp_path, media_file)
                            midia_reinserida += 1
                            self.log(f"Fallback ZIP: Re-inserido {os.path.basename(media_file)}.")

                shutil.rmtree(temp_dir)
                self.log(f"Fallback ZIP: {midia_reinserida} mídia(s) re-inseridas.")

        except Exception as zip_err:
            self.log(f"Aviso fallback ZIP: {zip_err}.")
            if os.path.exists('temp_media'):
                shutil.rmtree('temp_media')

        return midia_reinserida

    def calcular_valor_total(self, itens):
        """Calcula soma F13-F34 (formato BR)."""
        total_soma = 0.0
        itens_contados = 0
        try:
            for i, item in enumerate(itens):
                valor_total_str = item['total'].get_value()
                if valor_total_str and valor_total_str != "0,00":
                    valor_limpo = valor_total_str.replace('.', '').replace(',', '.')
                    try:
                        total_num = float(valor_limpo)
                        total_soma += total_num
                        itens_contados += 1
                    except ValueError:
                        self.log(f"Aviso: Total inválido item {i+1}: {valor_total_str}.")
            total_formatado = f"{total_soma:,.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
            self.log(f"F35 calculado: {total_formatado} (soma {itens_contados} itens).")
            return total_formatado
        except Exception as e:
            self.log(f"Erro cálculo F35: {e}")
            return "0,00"

    def aplicar_valores(self, ws_copia, proximo_numero):
        """Aplica valores na cópia (editados/reais; F35 calculado)."""
        try:
            # A5: Usa número atual da UI (baseado em A5 do modelo + incrementos)
            ws_copia['A5'] = f"N° do Orçamento {proximo_numero}"
            self.log(f"A5 aplicado na cópia: N° do Orçamento {proximo_numero} (baseado em modelo +1)")

            # Fixos B6-B9: só editados
            for campo in self.campos_fixos:
                valor = campo['entry'].get_value()
                if valor:
                    ws_copia[campo['celula_fixa']] = valor
                    self.log(f"Fixo '{campo['titulo']}' aplicado {campo['celula_fixa']}: {valor}")
                else:
                    ws_copia[campo['celula_fixa']] = ""
                    self.log(f"Fixo '{campo['titulo']}' em branco (não editado).")

            # Adicionais B36/B37/B39: real/editado
            for campo in self.campos_adicionais:
                valor = campo['entry'].get_value()
                if valor:
                    ws_copia[campo['celula']] = valor
                    self.log(f"Adicional '{campo['titulo']}' aplicado {campo['celula']}: {valor}")
                else:
                    ws_copia[campo['celula']] = ""
                    self.log(f"Adicional '{campo['titulo']}' em branco (não editado).")

            # Itens A13-F34: só editados
            for i, item in enumerate(self.itens_widgets):
                linha = 13 + i
                if linha > 34:
                    break
                colunas = [('item', 'A'), ('desc', 'B'), ('quant', 'C'), ('und', 'D'), ('vlr_uni', 'E'), ('total', 'F')]
                for key, col in colunas:
                    valor = item[key].get_value()
                    if valor:
                        celula = f"{col}{linha}"
                        ws_copia[celula] = valor
                        self.log(f"Item {i+1} {key.upper()} aplicado {celula}: {valor}")
                    else:
                        celula = f"{col}{linha}"
                        ws_copia[celula] = ""
                        self.log(f"Item {i+1} {key.upper()} em branco (não editado).")

            # F35: Calculado
            total_f35 = self.calcular_valor_total(self.itens_widgets)
            ws_copia['F35'] = total_f35
            self.log(f"F35 aplicado: {total_f35} (recalculado).")

            self.log("Valores aplicados (editados/reais; F35 calculado).")

        except Exception as e:
            self.log(f"Erro aplicar valores: {e}")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}")
            raise

    def gerar_pdf(self, caminho_arquivo):
        """Gera PDF via pywin32 (Windows)."""
        if not PDF_SUPORTE:
            self.log("PDF desabilitado (pywin32).")
            return None
        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(os.path.abspath(caminho_arquivo))
            ws = wb.ActiveSheet
            pasta_pdf = os.path.dirname(caminho_arquivo)
            nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
            pdf_caminho = os.path.join(pasta_pdf, f"{nome_base}.pdf")
            self.log(f"Gerando PDF: {pdf_caminho}")
            ws.ExportAsFixedFormat(0, pdf_caminho)  # 0 = PDF
            wb.Close(SaveChanges=False)
            excel.Quit()
            self.log("PDF gerado.")
            return pdf_caminho
        except Exception as e:
            self.log(f"Erro PDF: {e}")
            import traceback
            self.log(f"Traceback PDF: {traceback.format_exc()}")
            if 'excel' in locals():
                try:
                    excel.Quit()
                except:
                    pass
            return None

    def executar(self):
        """Executa criação (cópia, valores, imagens, PDF, persistência)."""
        self.botao_criar.config(state='disabled')
        self.botao_criar.configure(text="⏳ Processando...")
        self.root.update_idletasks()  # Loading visual
        self.log("Iniciando criação...")

        try:
            pasta = self.pasta_selecionada.get()
            if not pasta:
                raise ValueError("Selecione pasta.")
            arquivo_modelo = self.arquivo_selecionado.get()
            if not arquivo_modelo:
                raise ValueError("Selecione arquivo.")
            caminho_completo = os.path.join(pasta, arquivo_modelo)
            if not os.path.exists(caminho_completo):
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_completo}")

            proximo_numero = int(self.numero_orcamento.get())
            nome_base = os.path.splitext(arquivo_modelo)[0]
            novo_arquivo = f"{nome_base}_copia_{proximo_numero}.xlsx"
            novo_caminho = os.path.join(pasta, novo_arquivo)

            self.log(f"Copiando para {novo_arquivo}...")
            shutil.copy2(caminho_completo, novo_caminho)
            self.log("Arquivo copiado.")

            wb_copia = load_workbook(novo_caminho)
            ws_copia = wb_copia.active
            wb_original = load_workbook(caminho_completo)
            ws_original = wb_original.active

            self.copiar_imagens_completo(ws_original, ws_copia)

            self.aplicar_valores(ws_copia, proximo_numero)

            wb_copia.save(novo_caminho)
            self.log("Cópia salva.")

            imagens_detectadas = (hasattr(ws_original, '_images') and len(ws_original._images) > 0) or (hasattr(ws_original, '_drawings') and len(ws_original._drawings) > 0)
            if imagens_detectadas:
                self.fallback_zip_imagens(caminho_completo, novo_caminho)

            wb_original.close()

            self.log("Gerando PDF...")
            pdf_caminho = self.gerar_pdf(novo_caminho)

            # Persistência: +1 ao usado e salva JSON
            novo_proximo = proximo_numero + 1
            self.log(f"Incrementando A5: de {proximo_numero} para {novo_proximo}")
            self.numero_orcamento.set(str(novo_proximo))
            self.salvar_numero_config(novo_proximo)
            
            # Force update visual
            self.entry_numero.delete(0, tk.END)
            self.entry_numero.insert(0, str(novo_proximo))
            self.entry_numero.update_idletasks()
            print(f"DEBUG A5: UI atualizado para {novo_proximo} após +1")

            self.log(f"Cópia criada: {novo_arquivo} (A5: {proximo_numero}, F35 calculado, imagens OK).")
            if pdf_caminho:
                self.log(f"PDF: {os.path.basename(pdf_caminho)}")
                mensagem = f"✅ Cópia criada: {novo_arquivo}\n📄 PDF gerado: {os.path.basename(pdf_caminho)}\n📁 Pasta: {pasta}\n🔢 A5 usado: {proximo_numero} | Próximo: {novo_proximo}\n💾 Salvo em config.json"
            else:
                self.log("PDF não gerado (manual no Excel).")
                mensagem = f"✅ Cópia criada: {novo_arquivo}\n(Gere PDF manualmente no Excel)\n📁 Pasta: {pasta}\n🔢 A5 usado: {proximo_numero} | Próximo: {novo_proximo}\n💾 Salvo em config.json"

            self.log("Processo concluído com sucesso!")
            messagebox.showinfo("Sucesso!", mensagem)

            # App fica aberto para ver +1 (comente destroy se quiser fechar)
            # self.root.destroy()
            print("DEBUG A5: Verifique campo 'Número do Orçamento' agora em +1 (baseado em A5 do modelo)")

        except Exception as e:
            self.log(f"Erro execução: {e}")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}")
            if 'novo_caminho' in locals() and os.path.exists(novo_caminho):
                os.remove(novo_caminho)
                self.log("Cópia falha removida.")
            if 'pdf_caminho' in locals() and pdf_caminho and os.path.exists(pdf_caminho):
                os.remove(pdf_caminho)
                self.log("PDF falha removido.")
            messagebox.showerror("Erro!", f"Falha na criação: {e}\n\nLogs no terminal para detalhes.")
        finally:
            self.botao_criar.config(state='normal')
            self.botao_criar.configure(text="🚀 Criar Cópia do Orçamento e PDF" if PDF_SUPORTE else "💾 Criar Cópia do Orçamento (.xlsx apenas)")
            self.root.update_idletasks()
            self.log("Botão reabilitado. Pronto para nova operação.")

def main():
    """Função principal: Cria janela e inicia app."""
    root = tk.Tk()
    app = Automatizador(root)
    root.mainloop()

if __name__ == "__main__":
    main()
